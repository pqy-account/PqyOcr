import datetime

from openpyxl import load_workbook
from openpyxl.workbook import Workbook

# 加载工作簿
wb = load_workbook('C:\\Users\\Administrator\\Desktop\\2024发货数据 (1).xlsx')
ws = wb["每日更新数据"]

# 新的工作簿
wbNew = Workbook()
wsNew = wbNew.active

# 读取单元格数据
cell_value = ws['A1'].value
print(cell_value)

# 遍历工作表中的所有行
i = 0
for row in ws.iter_rows(values_only=True):
    if i == 0:
        print(row)
        i = i + 1
        newHeard = ("当单号长度", "仓库", "新日期",)
        heard = row + tuple(newHeard)
        wsNew.append(heard)
    else:
        cell_value = str(row[1])

        length = len(cell_value)
        excel_date_number = row[0]

        start_date = datetime.datetime(1899, 12, 30)  # Excel 从 1899 年 12 月 30 日开始计算
        date_value = start_date + datetime.timedelta(days=excel_date_number)

        if length >= 15:
            my_list = [length, "亚马逊", date_value]
            my_tuple = row + tuple(my_list)
            wsNew.append(my_tuple)
        else:
            my_list = [length, "独立站", date_value]
            my_tuple = row + tuple(my_list)
            wsNew.append(my_tuple)

wbNew.save("C:\\Users\\Administrator\\Desktop\\example.xlsx")
print("数据已写入到 example.xlsx 文件中.")
